"""
Context manager for write models.Goods to excel.
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import Goods


class XlsWriter:
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active

        headers = Goods.export_headers()
        self.ws.append(headers)

        self.max_widths = [len(str(h)) for h in headers]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self._apply_widths()
        self.wb.save(self.filename)

    def write_good(self, goods: Goods) -> None:
        row = goods.export_row()
        self.ws.append(row)

        for col_idx, value in enumerate(row):
            if value is None:
                continue

            text = str(value)

            length = max(len(line) for line in
                         text.split("\n")) if "\n" in text else len(text)

            if col_idx < len(self.max_widths):
                self.max_widths[col_idx] = max(self.max_widths[col_idx], length)

    def _apply_widths(self):
        for i, width in enumerate(self.max_widths, start=1):
            col_letter = get_column_letter(i)
            adjusted = min(width + 2, 120)
            self.ws.column_dimensions[col_letter].width = adjusted
